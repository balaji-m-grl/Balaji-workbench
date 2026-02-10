from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import sys




def compare_excel_sheets(old_file, new_file, output_file="Excel_Changes_Report.xlsx"):
    try:
        old_sheets = pd.read_excel(old_file, sheet_name=None, dtype=str)
        new_sheets = pd.read_excel(new_file, sheet_name=None, dtype=str)

        common_sheets = set(old_sheets.keys()) & set(new_sheets.keys())
        changes_detected = False

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet in common_sheets:
                old_df = old_sheets[sheet].fillna("").astype(str)
                new_df = new_sheets[sheet].fillna("").astype(str)

                # Maintain the column order from the original sheet
                column_order = list(old_df.columns)

                # Ensure all columns exist in both sheets
                for col in column_order:
                    if col not in new_df.columns:
                        new_df[col] = ""
                for col in new_df.columns:
                    if col not in old_df.columns:
                        old_df[col] = ""

                new_df = new_df[column_order]  # Ensure new_df follows the same column order as old_df
                old_df = old_df[column_order]  # Ensure old_df follows the same column order as old_df

                # Identify modifications first (before considering added/deleted rows)
                modified_rows = []
                for idx in old_df.index.intersection(new_df.index):
                    row_old, row_new = old_df.loc[idx], new_df.loc[idx]
                    if not row_old.equals(row_new):  # Check if any value is different
                        modified_row = {"Row Index": idx}
                        for col in column_order:
                            if row_old[col] != row_new[col]:  # Detect specific cell change
                                modified_row[col] = f"Old: {row_old[col]} → New: {row_new[col]}"
                        modified_rows.append(modified_row)

                # Identify Added and Deleted Rows
                new_rows = new_df.loc[~new_df.apply(tuple, axis=1).isin(old_df.apply(tuple, axis=1))]
                old_rows = old_df.loc[~old_df.apply(tuple, axis=1).isin(new_df.apply(tuple, axis=1))]

                # Exclude modified rows from added/deleted detection
                new_rows_filtered = new_rows.loc[~new_rows.index.isin(old_df.index)]
                old_rows_filtered = old_rows.loc[~old_rows.index.isin(new_df.index)]

                # If changes exist, write them into the report
                if modified_rows or not new_rows_filtered.empty or not old_rows_filtered.empty:
                    changes_detected = True
                    ws = writer.book.create_sheet(title=sheet)
                    row_count = 1

                    if modified_rows:
                        apply_formatting(ws, row_count, "MODIFIED ROWS", "FFF2CC")  # Yellow
                        pd.DataFrame(modified_rows).to_excel(writer, sheet_name=sheet, index=False, startrow=row_count)
                        row_count += len(modified_rows) + 3

                    if not new_rows_filtered.empty:
                        apply_formatting(ws, row_count, "ADDED ROWS", "E2F0D9")  # Green
                        new_rows_filtered.to_excel(writer, sheet_name=sheet, index=False, startrow=row_count)
                        row_count += len(new_rows_filtered) + 3

                    if not old_rows_filtered.empty:
                        apply_formatting(ws, row_count, "DELETED ROWS", "F4CCCC")  # Red
                        old_rows_filtered.to_excel(writer, sheet_name=sheet, index=False, startrow=row_count)

            # If no changes, create a simple summary
            if not changes_detected:
                pd.DataFrame({"Message": ["No changes detected"]}).to_excel(writer, sheet_name="Summary", index=False)

        apply_excel_highlight(output_file)

        print(f"📊 Comparison report saved to {output_file}")

    except Exception as e:
        print(f"❌ Error: {e}")

def apply_formatting(ws, row, section_name, color):
    """Apply section headers with bold black font and light background color."""
    cell = ws.cell(row=row, column=1, value=section_name)
    cell.font = Font(bold=True, color="000000")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def apply_excel_highlight(file_path):
    try:
        wb = load_workbook(file_path)

        light_green = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # Green - Added
        light_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow - Modified
        light_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Red - Deleted

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if "→" in str(cell.value):  # Highlight modified cells
                        cell.fill = light_yellow
                    elif str(cell.value).startswith("Old:"):
                        cell.fill = light_red  # Highlight old values
                    elif ws.cell(row=cell.row, column=1).value == "ADDED ROWS":
                        for c in ws[cell.row]:
                            c.fill = light_green
                    elif ws.cell(row=cell.row, column=1).value == "DELETED ROWS":
                        for c in ws[cell.row]:
                            c.fill = light_red

        wb.save(file_path)
        print("✅ Excel highlighting applied!")

    except Exception as e:
        print(f"⚠️ Failed to apply colors: {e}")
        
def excel_to_markdown(excel_file, output_md="Excel_Changes_Report.md"):
    sheets = pd.read_excel(excel_file, sheet_name=None, dtype=str)

    md_lines = []
    md_lines.append(f"# Excel Changes Report\n")

    for sheet_name, df in sheets.items():
        md_lines.append(f"\n## Sheet: {sheet_name}\n")

        if df.empty:
            md_lines.append("_No data in this sheet._\n")
            continue

        # Replace NaN with empty string
        df = df.fillna("")

        # Convert DataFrame to Markdown table
        md_table = df.to_markdown(index=False)

        md_lines.append(md_table)
        md_lines.append("\n")

    # Write to file
    output_path = Path(output_md)
    output_path.write_text("\n".join(md_lines), encoding="utf-8")

    print(f"✅ Markdown file created: {output_path.resolve()}")

if __name__ == "__main__":
    output_excel = "Excel_Changes_Report.xlsx"

    compare_excel_sheets(
        r"C:\Users\GRL\Downloads\timesheet.xlsx",
        r"C:\Users\GRL\Downloads\Balaji_BM-timesheet  new.xlsx",
        output_excel
    )

    excel_to_markdown(output_excel, "Excel_Changes_Report.md")